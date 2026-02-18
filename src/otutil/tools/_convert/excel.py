"""Excel workbook to Markdown converter.

Converts XLSX spreadsheets to Markdown with:
- Streaming row processing via openpyxl read_only mode
- Sheet-based sections
- Optional formula extraction
- YAML frontmatter and TOC generation
"""

from __future__ import annotations

from pathlib import Path  # noqa: TC003 (used at runtime)
from typing import Any

from otutil.tools._convert.utils import (
    IncrementalWriter,
    compute_file_checksum,
    get_mtime_iso,
    normalise_whitespace,
    write_toc_file,
)


def convert_excel(
    input_path: Path,
    output_dir: Path,
    source_rel: str,
    *,
    include_formulas: bool = False,
    compute_formulas: bool = False,
) -> dict[str, Any]:
    """Convert Excel workbook to Markdown.

    Args:
        input_path: Path to XLSX file
        output_dir: Directory for output files
        source_rel: Relative path to source for frontmatter
        include_formulas: Include cell formulas as comments
        compute_formulas: Evaluate formulas when cached values are missing
            (requires 'formulas' library: pip install formulas)

    Returns:
        Dict with 'output', 'sheets', 'rows' keys
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for convert. Install with: pip install openpyxl"
        ) from e

    # Load formula model if compute_formulas is enabled
    formula_model: Any = None
    if compute_formulas:
        try:
            import formulas  # type: ignore[import-untyped]

            formula_model = formulas.ExcelModel().loads(str(input_path)).finish()
        except ImportError:
            raise ImportError(
                "formulas library is required for compute_formulas. "
                "Install with: pip install formulas"
            ) from None
        except Exception:
            # If formula model fails to load, continue without it
            formula_model = None

    output_dir.mkdir(parents=True, exist_ok=True)

    # Load workbook
    # - read_only=False when computing formulas (need full access)
    # - data_only=True: get cached computed values (no formulas)
    # - data_only=False: get formulas as cell values (when include_formulas=True)
    read_only = not compute_formulas
    wb = load_workbook(input_path, read_only=read_only, data_only=not include_formulas)

    # Get metadata for frontmatter
    checksum = compute_file_checksum(input_path)
    mtime = get_mtime_iso(input_path)
    total_sheets = len(wb.sheetnames)

    writer = IncrementalWriter()
    total_rows = 0

    # Process each sheet (single workbook - no double loading)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = _process_sheet(writer, sheet_name, ws, include_formulas, formula_model)
        total_rows += rows

    wb.close()

    # Write main output (pure content, no frontmatter - line numbers start at 1)
    content = normalise_whitespace(writer.get_content())
    output_path = output_dir / f"{input_path.stem}.md"
    output_path.write_text(content, encoding="utf-8")

    # Write separate TOC file (includes frontmatter)
    headings = writer.get_headings()
    toc_path = write_toc_file(
        headings=headings,
        output_dir=output_dir,
        stem=input_path.stem,
        source=source_rel,
        converted=mtime,
        pages=total_sheets,
        checksum=checksum,
    )

    return {
        "output": str(output_path),
        "toc": str(toc_path),
        "sheets": total_sheets,
        "rows": total_rows,
    }


def _process_sheet(
    writer: IncrementalWriter,
    sheet_name: str,
    ws: Any,
    include_formulas: bool,
    formula_model: Any = None,
) -> int:
    """Process a single worksheet with streaming (O(1) memory for row data).

    When include_formulas=True, the workbook was loaded with data_only=False,
    so formula cells contain the formula string as their value.

    Args:
        writer: IncrementalWriter for output
        sheet_name: Name of the worksheet
        ws: Worksheet object
        include_formulas: Whether to include formulas in output
        formula_model: Optional formulas.ExcelModel for computing formula values

    Returns:
        Number of rows processed
    """
    writer.write_heading(2, f"Sheet: {sheet_name}")

    # First pass: count max columns (streaming, no data storage)
    max_cols = 0
    row_count = 0
    for row in ws.iter_rows():
        max_cols = max(max_cols, len(row))
        row_count += 1

    if row_count == 0:
        writer.write("(empty sheet)\n\n")
        return 0

    # Second pass: stream rows directly to writer
    rows_iter = iter(ws.iter_rows())

    # Get header (first row)
    first_row = next(rows_iter)
    header = [
        _get_cell_value(cell, sheet_name, 1, j + 1, formula_model)
        for j, cell in enumerate(first_row)
    ]
    # Pad header to max_cols
    while len(header) < max_cols:
        header.append("")

    # Write header
    writer.write("| " + " | ".join(_escape_pipe(c) for c in header) + " |\n")
    writer.write("| " + " | ".join("---" for _ in header) + " |\n")

    # Collect formulas as we go (just formula tuples, not full row data)
    # Format: (col_letter, row_num, formula_string)
    formulas: list[tuple[str, int, str]] = []

    # Check first row for formulas (cell values are formulas when include_formulas=True)
    if include_formulas:
        for j, cell in enumerate(first_row):
            try:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas.append((_col_letter(j + 1), 1, value))
            except Exception:
                pass

    # Stream remaining rows directly to writer
    current_row = 2  # 1-indexed, header was row 1
    for row in rows_iter:
        row_values = [
            _get_cell_value(cell, sheet_name, current_row, j + 1, formula_model)
            for j, cell in enumerate(row)
        ]
        # Pad row to max_cols
        while len(row_values) < max_cols:
            row_values.append("")

        writer.write("| " + " | ".join(_escape_pipe(c) for c in row_values[:len(header)]) + " |\n")

        # Track formulas for this row (cell values are formulas when include_formulas=True)
        if include_formulas:
            for j, cell in enumerate(row):
                try:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formulas.append((_col_letter(j + 1), current_row, value))
                except Exception:
                    pass

        current_row += 1

    writer.write("\n")

    # Add formulas section if any formulas found
    if formulas:
        writer.write("**Formulas:**\n\n")
        writer.write("```\n")
        for col_letter, row_num, formula in formulas:
            writer.write(f"{col_letter}{row_num}: {formula}\n")
        writer.write("```\n\n")

    return row_count


def _get_cell_value(
    cell: Any,
    sheet_name: str,
    row_num: int,
    col_num: int,
    formula_model: Any,
) -> str:
    """Get cell value, optionally computing from formula model.

    Args:
        cell: openpyxl cell object
        sheet_name: Name of the worksheet (for formula lookup)
        row_num: 1-indexed row number
        col_num: 1-indexed column number
        formula_model: Optional formulas.ExcelModel for computing values

    Returns:
        String representation of cell value
    """
    value = cell.value

    # If we have a value, use it
    if value is not None:
        return str(value)

    # If no formula model, return empty
    if formula_model is None:
        return ""

    # Try to compute value from formula model
    try:
        # Build cell reference like "'Sheet1'!A1"
        col_letter = _col_letter(col_num)
        cell_ref = f"'{sheet_name}'!{col_letter}{row_num}"
        computed = formula_model.calculate(cell_ref)
        if computed is not None and computed != cell_ref:
            # Handle numpy arrays and other types
            if hasattr(computed, "item"):
                computed = computed.item()
            return str(computed)
    except Exception:
        pass

    return ""


def _escape_pipe(text: str) -> str:
    """Escape pipe characters for Markdown tables."""
    return text.replace("|", "\\|").replace("\n", " ")


def _col_letter(n: int) -> str:
    """Convert column number to letter (1=A, 2=B, ..., 27=AA)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result
