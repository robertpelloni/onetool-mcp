"""Excel file manipulation tools.

Create, read, write Excel workbooks using openpyxl.

Based on excel-mcp-server by Haris Musa (MIT License).
https://github.com/haris-musa/excel-mcp-server
"""

from __future__ import annotations

# Pack for dot notation: excel.create(), excel.read(), etc.
pack = "excel"

__all__ = [
    "add_sheet",
    "cell_range",
    "cell_shift",
    "copy_range",
    "create",
    "create_table",
    "delete_cols",
    "delete_rows",
    "formula",
    "formulas",
    "hyperlinks",
    "info",
    "insert_cols",
    "insert_rows",
    "merged_cells",
    "named_ranges",
    "read",
    "search",
    "sheets",
    "table_data",
    "table_info",
    "tables",
    "used_range",
    "write",
]

# Dependency declarations for CLI validation
__ot_requires__ = {
    "lib": [("openpyxl", "pip install openpyxl")],
}

import fnmatch
import re
from datetime import date, datetime, time
from typing import TYPE_CHECKING, Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
    from openpyxl.worksheet.cell_range import CellRange
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as _openpyxl_err:
    raise ImportError(
        "Excel tools require the [util] extra. "
        "Install with: pip install onetool-mcp[util]"
    ) from _openpyxl_err
from otpack import LogSpan, get_tool_config, resolve_cwd_path
from pydantic import BaseModel

if TYPE_CHECKING:
    from pathlib import Path


class Config(BaseModel):
    """Pack configuration - discovered by registry."""

    pass


def _get_config() -> Config:
    """Get the tool configuration."""
    return get_tool_config("excel", Config)


def _expand_path(filepath: str) -> Path:
    """Resolve a file path relative to project directory.

    Uses SDK resolve_cwd_path() for consistent path resolution.

    Path resolution follows project conventions:
        - Relative paths: resolved relative to project directory (OT_CWD)
        - Absolute paths: used as-is
        - ~ paths: expanded to home directory
        - Prefixed paths (CWD/, GLOBAL/, OT_DIR/): resolved to respective dirs

    Note: ${VAR} patterns are NOT expanded. Use ~/path instead of ${HOME}/path.

    Args:
        filepath: Path string (can contain ~ or prefixes)

    Returns:
        Resolved absolute Path
    """
    return resolve_cwd_path(filepath)


def _ensure_parent_dir(filepath: str) -> None:
    """Create parent directories if they don't exist."""
    parent = _expand_path(filepath).parent
    if parent and not parent.exists():
        parent.mkdir(parents=True, exist_ok=True)


def _get_sheet(wb: Workbook, sheet_name: str | None) -> tuple[Any, str | None]:
    """Get worksheet by name or return active sheet.

    Returns:
        Tuple of (worksheet, error_message). On success, error is None.
        On failure, worksheet is None and error contains the message.
    """
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return (
                None,
                f"Error: Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}",
            )
        return wb[sheet_name], None
    return wb.active, None


def _col_to_index(col: int | str) -> int:
    """Convert column letter or number to 1-based index."""
    if isinstance(col, int):
        return col
    return column_index_from_string(col)


def _plural(count: int, singular: str, plural: str | None = None) -> str:
    """Return singular or plural form based on count."""
    if count == 1:
        return singular
    return plural or f"{singular}s"


def create(
    *,
    filepath: str,
    sheet_name: str = "Sheet1",
    sheet_names: list[str] | None = None,
) -> str:
    """Create new Excel workbook.

    Args:
        filepath: Path to create the Excel file
        sheet_name: Name for the initial sheet (default: "Sheet1"); ignored when sheet_names is given
        sheet_names: Create multiple sheets in one call (e.g., ["Sales", "Config", "Summary"])

    Returns:
        Success message with filepath

    Example:
        excel.create(filepath="output/report.xlsx")
        excel.create(filepath="data.xlsx", sheet_name="Sales")
        excel.create(filepath="data.xlsx", sheet_names=["Sales", "Config", "Summary"])
    """
    with LogSpan(span="excel.create", filepath=filepath, sheet=sheet_name) as s:
        try:
            _ensure_parent_dir(filepath)
            wb = Workbook()
            ws = wb.active
            if sheet_names:
                ws.title = sheet_names[0]
                for name in sheet_names[1:]:
                    wb.create_sheet(name)
            else:
                ws.title = sheet_name
            wb.save(_expand_path(filepath))
            s.add(created=True)
            return f"Created workbook: {filepath}"
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def add_sheet(*, filepath: str, sheet_name: str) -> str:
    """Add worksheet to existing workbook.

    Args:
        filepath: Path to Excel file
        sheet_name: Name for the new sheet

    Returns:
        Success message or error if sheet exists

    Example:
        excel.add_sheet(filepath="report.xlsx", sheet_name="Summary")
    """
    with LogSpan(span="excel.add_sheet", filepath=filepath, sheet=sheet_name) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            if sheet_name in wb.sheetnames:
                s.add(error="sheet_exists")
                return f"Error: Sheet '{sheet_name}' already exists"

            wb.create_sheet(title=sheet_name)
            wb.save(_expand_path(filepath))
            s.add(created=True)
            return f"Created sheet: {sheet_name}"
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def read(
    *,
    filepath: str,
    sheet_name: str | None = None,
    start_cell: str = "A1",
    end_cell: str | None = None,
) -> list[list[Any]] | str:
    """Read data from Excel worksheet.

    Args:
        filepath: Path to Excel file
        sheet_name: Sheet to read (default: active sheet)
        start_cell: Starting cell reference (default: "A1")
        end_cell: Ending cell reference (default: auto-detect)

    Returns:
        List of rows (each a list of cell values), or error string

    Example:
        excel.read(filepath="data.xlsx")
        excel.read(filepath="data.xlsx", sheet_name="Sales", start_cell="B2", end_cell="D10")
    """
    with LogSpan(span="excel.read", filepath=filepath, sheet=sheet_name) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath), data_only=True)
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            # Determine range
            if end_cell:
                cell_range = f"{start_cell}:{end_cell}"
            else:
                # Auto-detect used range
                # openpyxl reports max_row=1, max_column=1 even for empty sheets,
                # so check A1 value to distinguish truly empty from single-cell data.
                empty = (
                    ws.max_row == 1
                    and ws.max_column == 1
                    and ws["A1"].value is None
                )
                if empty:
                    s.add(rows=0)
                    return "No data in worksheet"
                end_col = get_column_letter(ws.max_column)
                cell_range = f"{start_cell}:{end_col}{ws.max_row}"

            # Read data
            rows = []
            for row in ws[cell_range]:
                row_data = []
                for cell in row:
                    value = cell.value
                    if value is None:
                        row_data.append("")
                    elif isinstance(value, (datetime, date, time)):
                        row_data.append(value.isoformat())
                    else:
                        row_data.append(value)
                rows.append(row_data)

            s.add(rows=len(rows))
            return rows
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def write(
    *,
    filepath: str,
    data: list[list[Any]],
    sheet_name: str | None = None,
    start_cell: str = "A1",
    create_if_missing: bool = False,
) -> str:
    """Write data to Excel worksheet.

    Args:
        filepath: Path to Excel file
        data: List of rows, where each row is a list of values
        sheet_name: Sheet to write to (default: active sheet)
        start_cell: Starting cell reference (default: "A1")
        create_if_missing: Create file if it doesn't exist (default: False)

    Returns:
        Success message with row count

    Example:
        excel.write(filepath="report.xlsx", data=[["Name", "Score"], ["Alice", 95]])
        excel.write(filepath="report.xlsx", data=[[1, 2, 3]], sheet_name="Numbers", start_cell="B5")
        excel.write(filepath="new.xlsx", data=[["Test"]], create_if_missing=True)
    """
    with LogSpan(span="excel.write", filepath=filepath, sheet=sheet_name, rows=len(data)) as s:
        try:
            path = _expand_path(filepath)
            if not path.exists():
                if not create_if_missing:
                    s.add(error="file_not_found")
                    return f"Error: File not found: {filepath}"
                _ensure_parent_dir(filepath)
                wb = Workbook()
                if sheet_name:
                    wb.active.title = sheet_name
            else:
                wb = load_workbook(path)

            if not data:
                s.add(error="no_data")
                return "Error: No data provided"
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            # Parse start cell
            col_letter, start_row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(col_letter)

            # Write data
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    ws.cell(
                        row=start_row + row_idx,
                        column=start_col + col_idx,
                        value=value,
                    )

            wb.save(_expand_path(filepath))
            sheet_used = sheet_name or ws.title
            s.add(written=True)
            return f"Wrote {len(data)} {_plural(len(data), 'row')} to {sheet_used}"
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def info(*, filepath: str, include_ranges: bool = False) -> dict[str, Any] | str:
    """Get workbook metadata.

    Args:
        filepath: Path to Excel file
        include_ranges: Include used range for each sheet (default: False)

    Returns:
        Formatted info with filename, sheets, size, and optionally ranges

    Example:
        excel.info(filepath="report.xlsx")
        excel.info(filepath="data.xlsx", include_ranges=True)
    """
    with LogSpan(span="excel.info", filepath=filepath) as s:
        try:
            resolved = _expand_path(filepath)
            if not resolved.exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            # Use read_only=True only when we don't need ranges
            # (read_only mode doesn't populate max_row/max_column accurately)
            wb = load_workbook(resolved, read_only=not include_ranges)
            file_size = resolved.stat().st_size

            info_dict: dict[str, Any] = {
                "file": resolved.name,
                "sheets": wb.sheetnames,
                "size": f"{file_size:,} bytes",
            }

            if include_ranges:
                ranges = {}
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if ws.max_row and ws.max_column:
                        end_col = get_column_letter(ws.max_column)
                        ranges[sheet_name] = f"A1:{end_col}{ws.max_row}"
                    else:
                        ranges[sheet_name] = "empty"
                info_dict["ranges"] = ranges

            wb.close()
            s.add(sheets=len(info_dict["sheets"]))
            return info_dict
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def formula(
    *,
    filepath: str,
    cell: str,
    formula: str,
    sheet_name: str | None = None,
) -> str:
    """Apply Excel formula to a cell.

    Args:
        filepath: Path to Excel file
        cell: Cell reference (e.g., "A1", "B10")
        formula: Excel formula (= prefix added automatically if missing)
        sheet_name: Sheet name (default: active sheet)

    Returns:
        Success message with applied formula

    Example:
        excel.formula(filepath="sales.xlsx", cell="C10", formula="=SUM(C2:C9)")
        excel.formula(filepath="data.xlsx", cell="A1", formula="=TODAY()", sheet_name="Summary")
    """
    with LogSpan(span="excel.formula", filepath=filepath, cell=cell) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            # Auto-prepend = if missing
            formula_str = formula if formula.startswith("=") else f"={formula}"

            ws[cell] = formula_str
            wb.save(_expand_path(filepath))
            s.add(applied=True)
            return f"Applied formula to {cell}: {formula_str}"
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


# =============================================================================
# Tier 1: Range Manipulation (Pure Functions)
# =============================================================================


def cell_range(
    *,
    cell: str,
    right: int = 0,
    down: int = 0,
    left: int = 0,
    up: int = 0,
) -> str:
    """(no file) Expand a cell into a range using CellRange.expand().

    Pure function - no file required.

    Args:
        cell: Starting cell reference (e.g., "A1")
        right: Expand right by N columns
        down: Expand down by N rows
        left: Expand left by N columns
        up: Expand up by N rows

    Returns:
        Range reference (e.g., "A1:F6")

    Example:
        excel.cell_range(cell="A1", right=5, down=5)  # -> "A1:F6"
        excel.cell_range(cell="C3", left=2, up=2)     # -> "A1:C3"
    """
    with LogSpan(span="excel.cell_range", cell=cell, right=right, down=down) as s:
        try:
            r = CellRange(cell)
            r.expand(right=right, down=down, left=left, up=up)
            result = str(r.coord)
            s.add(result=result)
            return result
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def cell_shift(
    *,
    cell: str,
    rows: int = 0,
    cols: int = 0,
) -> str:
    """(no file) Shift a cell reference using CellRange.shift().

    Pure function - no file required.

    Args:
        cell: Starting cell reference (e.g., "A1")
        rows: Rows to shift (positive=down, negative=up)
        cols: Columns to shift (positive=right, negative=left)

    Returns:
        New cell reference

    Example:
        excel.cell_shift(cell="A1", rows=5)         # -> "A6"
        excel.cell_shift(cell="A1", cols=5)         # -> "F1"
        excel.cell_shift(cell="B3", rows=2, cols=3) # -> "E5"
    """
    with LogSpan(span="excel.cell_shift", cell=cell, rows=rows, cols=cols) as s:
        try:
            r = CellRange(cell)
            r.shift(row_shift=rows, col_shift=cols)
            result = str(r.coord)
            s.add(result=result)
            return result
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


# =============================================================================
# Tier 1: Search
# =============================================================================


def search(
    *,
    filepath: str,
    pattern: str,
    sheet_name: str | None = None,
    regex: bool = False,
    first_only: bool = False,
) -> list[dict[str, str]] | str:
    """Search for values matching a pattern.

    Args:
        filepath: Path to Excel file
        pattern: Search pattern (wildcards * ? if not regex)
        sheet_name: Sheet to search (default: active sheet)
        regex: Treat pattern as regex (default: False)
        first_only: Return only first match (default: False)

    Returns:
        JSON list of matches: [{cell: "A1", value: "found text"}, ...]

    Example:
        excel.search(filepath="data.xlsx", pattern="Error*")
        excel.search(filepath="data.xlsx", pattern="^ID-\\\\d+$", regex=True)
        excel.search(filepath="data.xlsx", pattern="Total", first_only=True)
    """
    with LogSpan(span="excel.search", filepath=filepath, pattern=pattern) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath), data_only=True)
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            matches: list[dict[str, str]] = []
            compiled_regex = re.compile(pattern) if regex else None

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    matched = False

                    if regex:
                        if compiled_regex and compiled_regex.search(text):
                            matched = True
                    else:
                        if fnmatch.fnmatch(text, pattern):
                            matched = True

                    if matched:
                        matches.append({"cell": cell.coordinate, "value": text})
                        if first_only:
                            s.add(resultCount=1)
                            return [matches[0]]

            wb.close()
            s.add(resultCount=len(matches))
            return matches
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


# =============================================================================
# Tier 1: Table Access
# =============================================================================


def tables(
    *,
    filepath: str,
    sheet_name: str | None = None,
) -> list[dict[str, str]] | str:
    """List all defined tables in worksheet.

    Args:
        filepath: Path to Excel file
        sheet_name: Sheet to inspect (default: active sheet)

    Returns:
        JSON list of table info: [{name, ref}, ...]

    Example:
        excel.tables(filepath="sales.xlsx")
    """
    with LogSpan(span="excel.tables", filepath=filepath) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            table_list = [
                {"name": table.name, "ref": table.ref} for table in ws.tables.values()
            ]

            wb.close()
            s.add(resultCount=len(table_list))
            return table_list
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def table_info(
    *,
    filepath: str,
    table_name: str,
    sheet_name: str | None = None,
) -> dict[str, Any] | str:
    """Get detailed table information.

    Args:
        filepath: Path to Excel file
        table_name: Name of the table
        sheet_name: Sheet containing table (default: active sheet)

    Returns:
        JSON dict: name, ref, headers, row_count, has_totals

    Example:
        excel.table_info(filepath="sales.xlsx", table_name="SalesData")
    """
    with LogSpan(span="excel.table_info", filepath=filepath, table=table_name) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            if table_name not in ws.tables:
                s.add(error="table_not_found")
                wb.close()
                return f"Error: Table '{table_name}' not found"

            table = ws.tables[table_name]
            # Parse ref to get row count
            ref_range = CellRange(table.ref)
            data_rows = ref_range.max_row - ref_range.min_row  # Excludes header

            info_dict = {
                "name": table.name,
                "ref": table.ref,
                "headers": list(table.column_names) if table.column_names else [],
                "row_count": data_rows,
                "has_totals": table.totalsRowCount > 0
                if table.totalsRowCount
                else False,
            }

            wb.close()
            s.add(found=True)
            return info_dict
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def table_data(
    *,
    filepath: str,
    table_name: str,
    row_index: int | None = None,
    sheet_name: str | None = None,
) -> dict[str, Any] | list[dict[str, Any]] | str:
    """Get table data with optional row selection.

    Args:
        filepath: Path to Excel file
        table_name: Name of the table
        row_index: Specific row (0-indexed, excludes header). None = all rows
        sheet_name: Sheet containing table

    Returns:
        Single row: JSON dict {header: value, ...}
        All rows: JSON list of dicts

    Example:
        excel.table_data(filepath="sales.xlsx", table_name="SalesData")
        excel.table_data(filepath="sales.xlsx", table_name="SalesData", row_index=0)
    """
    with LogSpan(span="excel.table_data", filepath=filepath, table=table_name) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath), data_only=True)
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            if table_name not in ws.tables:
                s.add(error="table_not_found")
                wb.close()
                return f"Error: Table '{table_name}' not found"

            table = ws.tables[table_name]
            headers = list(table.column_names) if table.column_names else []
            ref_range = CellRange(table.ref)

            # Read data rows (skip header row)
            rows_data = []
            for row in ws.iter_rows(
                min_row=ref_range.min_row + 1,
                max_row=ref_range.max_row,
                min_col=ref_range.min_col,
                max_col=ref_range.max_col,
            ):
                row_dict = {}
                for col_idx, cell in enumerate(row):
                    header = (
                        headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
                    )
                    v = cell.value
                    if v is None:
                        row_dict[header] = ""
                    elif isinstance(v, (datetime, date, time)):
                        row_dict[header] = v.isoformat()
                    else:
                        row_dict[header] = v
                rows_data.append(row_dict)

            wb.close()

            if row_index is not None:
                if 0 <= row_index < len(rows_data):
                    s.add(resultCount=1)
                    return rows_data[row_index]
                else:
                    s.add(error="row_index_out_of_range")
                    return f"Error: Row index {row_index} out of range (0-{len(rows_data) - 1})"

            s.add(resultCount=len(rows_data))
            return rows_data
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


# =============================================================================
# Tier 2: Structure Manipulation
# =============================================================================


def insert_rows(
    *,
    filepath: str,
    row: int,
    count: int = 1,
    sheet_name: str | None = None,
) -> str:
    """Insert rows at specified position.

    Args:
        filepath: Path to Excel file
        row: Row number to insert at (1-based)
        count: Number of rows to insert (default: 1)
        sheet_name: Sheet to modify (default: active sheet)

    Returns:
        Success message

    Example:
        excel.insert_rows(filepath="data.xlsx", row=5, count=3)
    """
    with LogSpan(span="excel.insert_rows", filepath=filepath, row=row, count=count) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            ws.insert_rows(row, count)
            wb.save(_expand_path(filepath))
            wb.close()
            s.add(inserted=count)
            return f"Inserted {count} {_plural(count, 'row')} at row {row}"
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def delete_rows(
    *,
    filepath: str,
    row: int,
    count: int = 1,
    sheet_name: str | None = None,
) -> str:
    """Delete rows starting at specified position.

    Args:
        filepath: Path to Excel file
        row: Row number to start deleting (1-based)
        count: Number of rows to delete (default: 1)
        sheet_name: Sheet to modify (default: active sheet)

    Returns:
        Success message

    Example:
        excel.delete_rows(filepath="data.xlsx", row=3, count=2)
    """
    with LogSpan(span="excel.delete_rows", filepath=filepath, row=row, count=count) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            ws.delete_rows(row, count)
            wb.save(_expand_path(filepath))
            wb.close()
            s.add(deleted=count)
            return f"Deleted {count} {_plural(count, 'row')} starting at row {row}"
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def insert_cols(
    *,
    filepath: str,
    col: int | str,
    count: int = 1,
    sheet_name: str | None = None,
) -> str:
    """Insert columns at specified position.

    Args:
        filepath: Path to Excel file
        col: Column number (1-based) or letter ("A", "B", etc.)
        count: Number of columns to insert (default: 1)
        sheet_name: Sheet to modify (default: active sheet)

    Returns:
        Success message

    Example:
        excel.insert_cols(filepath="data.xlsx", col="C", count=2)
        excel.insert_cols(filepath="data.xlsx", col=3, count=2)
    """
    with LogSpan(span="excel.insert_cols", filepath=filepath, col=col, count=count) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            col_idx = _col_to_index(col)
            col_letter = get_column_letter(col_idx)
            ws.insert_cols(col_idx, count)
            wb.save(_expand_path(filepath))
            wb.close()
            s.add(inserted=count)
            return f"Inserted {count} {_plural(count, 'column')} at column {col_letter}"
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def delete_cols(
    *,
    filepath: str,
    col: int | str,
    count: int = 1,
    sheet_name: str | None = None,
) -> str:
    """Delete columns starting at specified position.

    Args:
        filepath: Path to Excel file
        col: Column number (1-based) or letter ("A", "B", etc.)
        count: Number of columns to delete (default: 1)
        sheet_name: Sheet to modify (default: active sheet)

    Returns:
        Success message

    Example:
        excel.delete_cols(filepath="data.xlsx", col="B", count=2)
    """
    with LogSpan(span="excel.delete_cols", filepath=filepath, col=col, count=count) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            col_idx = _col_to_index(col)
            col_letter = get_column_letter(col_idx)
            ws.delete_cols(col_idx, count)
            wb.save(_expand_path(filepath))
            wb.close()
            s.add(deleted=count)
            return f"Deleted {count} {_plural(count, 'column')} starting at column {col_letter}"
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def copy_range(
    *,
    filepath: str,
    source_range: str,
    target_cell: str,
    sheet_name: str | None = None,
    target_sheet: str | None = None,
) -> str:
    """Copy a range to another location.

    Args:
        filepath: Path to Excel file
        source_range: Source range (e.g., "A1:C10")
        target_cell: Target cell (top-left of destination)
        sheet_name: Source sheet (default: active sheet)
        target_sheet: Target sheet (default: same as source)

    Returns:
        Success message

    Example:
        excel.copy_range(filepath="data.xlsx", source_range="A1:C10", target_cell="E1")
        excel.copy_range(filepath="data.xlsx", source_range="A1:C10", target_cell="A1", target_sheet="Backup")
    """
    with LogSpan(span="excel.copy_range", filepath=filepath, source=source_range, target=target_cell) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws_source, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            # Get target worksheet
            if target_sheet:
                ws_target, err = _get_sheet(wb, target_sheet)
                if err:
                    wb.close()
                    s.add(error="target_sheet_not_found")
                    return err
            else:
                ws_target = ws_source

            # Parse source range
            src_range = CellRange(source_range)
            # Parse target cell
            target_col_letter, target_row = coordinate_from_string(target_cell)
            target_col = column_index_from_string(target_col_letter)

            # Copy cells
            for row_offset, row in enumerate(
                ws_source.iter_rows(
                    min_row=src_range.min_row,
                    max_row=src_range.max_row,
                    min_col=src_range.min_col,
                    max_col=src_range.max_col,
                )
            ):
                for col_offset, cell in enumerate(row):
                    ws_target.cell(
                        row=target_row + row_offset,
                        column=target_col + col_offset,
                        value=cell.value,
                    )

            # Calculate destination range for message
            dest_end_col = get_column_letter(
                target_col + src_range.max_col - src_range.min_col
            )
            dest_end_row = target_row + src_range.max_row - src_range.min_row
            dest_range = f"{target_cell}:{dest_end_col}{dest_end_row}"

            wb.save(_expand_path(filepath))
            wb.close()
            s.add(copied=True)
            return f"Copied {source_range} to {dest_range}"
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def create_table(
    *,
    filepath: str,
    data_range: str,
    table_name: str | None = None,
    sheet_name: str | None = None,
) -> str:
    """Create a native Excel table from a data range.

    Tables enable filtering, sorting, and structured references.
    First row of range is used as headers.

    Args:
        filepath: Path to Excel file
        data_range: Range containing data (e.g., "A1:E10")
        table_name: Name for the table (default: auto-generated)
        sheet_name: Sheet containing data (default: active sheet)

    Returns:
        Success message

    Example:
        excel.create_table(filepath="sales.xlsx", data_range="A1:E10", table_name="SalesData")
    """
    with LogSpan(span="excel.create_table", filepath=filepath, range=data_range) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            # Generate table name if not provided
            if table_name is None:
                existing_tables = set(ws.tables.keys())
                counter = 1
                while f"Table{counter}" in existing_tables:
                    counter += 1
                table_name = f"Table{counter}"

            # Create table with default style
            table = Table(displayName=table_name, ref=data_range)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

            wb.save(_expand_path(filepath))
            wb.close()
            s.add(created=table_name)
            return f"Created table '{table_name}' from {data_range}"
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


# =============================================================================
# Tier 3: Extended Inspection
# =============================================================================


def sheets(*, filepath: str) -> list[dict[str, str]] | str:
    """List all sheets with visibility and type.

    Args:
        filepath: Path to Excel file

    Returns:
        JSON list: [{name, state}, ...]
        state: 'visible', 'hidden', 'veryHidden'

    Example:
        excel.sheets(filepath="report.xlsx")
    """
    with LogSpan(span="excel.sheets", filepath=filepath) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            sheet_list = []
            for name in wb.sheetnames:
                ws = wb[name]
                state = ws.sheet_state if ws.sheet_state else "visible"
                sheet_list.append({"name": name, "state": state})

            wb.close()
            s.add(resultCount=len(sheet_list))
            return sheet_list
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def used_range(
    *,
    filepath: str,
    sheet_name: str | None = None,
) -> str:
    """Get the used range of a worksheet.

    Args:
        filepath: Path to Excel file
        sheet_name: Sheet to inspect (default: active sheet)

    Returns:
        Range reference (e.g., "A1:Z100") or "empty"

    Example:
        excel.used_range(filepath="data.xlsx")
    """
    with LogSpan(span="excel.used_range", filepath=filepath) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            if ws.max_row and ws.max_column:
                end_col = get_column_letter(ws.max_column)
                result = f"A1:{end_col}{ws.max_row}"
            else:
                result = "empty"

            wb.close()
            s.add(result=result)
            return result
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def formulas(
    *,
    filepath: str,
    sheet_name: str | None = None,
) -> list[dict[str, str]] | str:
    """List all cells containing formulas.

    Args:
        filepath: Path to Excel file
        sheet_name: Sheet to inspect (default: active sheet)

    Returns:
        JSON list: [{cell, formula}, ...]

    Example:
        excel.formulas(filepath="calc.xlsx")
    """
    with LogSpan(span="excel.formulas", filepath=filepath) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            # Don't use data_only to preserve formulas
            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            formula_list = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or (
                        isinstance(cell.value, str) and cell.value.startswith("=")
                    ):
                        formula_list.append(
                            {
                                "cell": cell.coordinate,
                                "formula": cell.value,
                            }
                        )

            wb.close()
            s.add(resultCount=len(formula_list))
            return formula_list
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def hyperlinks(
    *,
    filepath: str,
    sheet_name: str | None = None,
) -> list[dict[str, str]] | str:
    """List all hyperlinks in worksheet.

    Args:
        filepath: Path to Excel file
        sheet_name: Sheet to inspect (default: active sheet)

    Returns:
        JSON list: [{cell, target, display}, ...]

    Example:
        excel.hyperlinks(filepath="links.xlsx")
    """
    with LogSpan(span="excel.hyperlinks", filepath=filepath) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            link_list = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        link_list.append(
                            {
                                "cell": cell.coordinate,
                                "target": cell.hyperlink.target or "",
                                "display": str(cell.value) if cell.value else "",
                            }
                        )

            wb.close()
            s.add(resultCount=len(link_list))
            return link_list
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def merged_cells(
    *,
    filepath: str,
    sheet_name: str | None = None,
) -> list[str] | str:
    """List merged cell ranges in worksheet.

    Args:
        filepath: Path to Excel file
        sheet_name: Sheet to inspect (default: active sheet)

    Returns:
        JSON list of range strings: ["B2:F4", "A10:C10", ...]

    Example:
        excel.merged_cells(filepath="report.xlsx")
    """
    with LogSpan(span="excel.merged_cells", filepath=filepath) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            ws, err = _get_sheet(wb, sheet_name)
            if err:
                wb.close()
                s.add(error="sheet_not_found")
                return err

            merged_list = [str(r) for r in ws.merged_cells.ranges]

            wb.close()
            s.add(resultCount=len(merged_list))
            return merged_list
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"


def named_ranges(*, filepath: str) -> list[dict[str, Any]] | str:
    """List all named ranges in workbook.

    Args:
        filepath: Path to Excel file

    Returns:
        JSON list: [{name, value, destinations}, ...]

    Example:
        excel.named_ranges(filepath="report.xlsx")
    """
    with LogSpan(span="excel.named_ranges", filepath=filepath) as s:
        try:
            if not _expand_path(filepath).exists():
                s.add(error="file_not_found")
                return f"Error: File not found: {filepath}"

            wb = load_workbook(_expand_path(filepath))
            range_list = []

            for defn in wb.defined_names.values():
                destinations = []
                try:
                    for sheet_title, cell_range in defn.destinations:
                        destinations.append(f"{sheet_title}!{cell_range}")
                except Exception:
                    pass

                range_list.append(
                    {
                        "name": defn.name,
                        "value": defn.attr_text,
                        "destinations": destinations,
                    }
                )

            wb.close()
            s.add(resultCount=len(range_list))
            return range_list
        except Exception as e:
            s.add(error=str(e))
            return f"Error: {e}"
