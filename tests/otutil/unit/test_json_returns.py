"""Tests verifying structured-return tools produce valid JSON strings (task 4.4)."""

from __future__ import annotations

import json
import os
from pathlib import Path

import pytest


@pytest.fixture
def tmp_file(tmp_path: Path) -> Path:
    """Create a temp file for testing."""
    f = tmp_path / "test.txt"
    f.write_text("hello world")
    return f


@pytest.fixture
def tmp_excel(tmp_path: Path) -> Path:
    """Create a temp Excel file for testing."""
    pytest.importorskip("openpyxl", reason="openpyxl not installed")
    from openpyxl import Workbook

    f = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "foo"
    ws["B2"] = 42
    wb.save(str(f))
    wb.close()
    return f


@pytest.mark.unit
@pytest.mark.tools
class TestFileJsonReturns:
    """file.info() returns structured data."""

    def test_info_returns_dict(self, tmp_file: Path) -> None:
        """file.info returns a dict, not a JSON string."""
        from otutil.tools.file import info

        # Run from tmp_path dir so path validation allows it
        old_cwd = os.getcwd()
        try:
            os.chdir(tmp_file.parent)
            result = info(path=str(tmp_file))
        finally:
            os.chdir(old_cwd)

        assert isinstance(result, dict), f"Expected dict, got {type(result)}"
        assert "path" in result
        assert "size" in result
        assert "type" in result


@pytest.mark.unit
@pytest.mark.tools
class TestExcelJsonReturns:
    """Excel tools return native Python types (serialized by serialize_result())."""

    def test_read_returns_list(self, tmp_excel: Path) -> None:
        """excel.read returns a list of rows."""
        from otutil.tools.excel import read

        result = read(filepath=str(tmp_excel))
        assert isinstance(result, list), f"Expected list, got {type(result)}"
        assert len(result) >= 2  # header + data row

    def test_info_returns_dict(self, tmp_excel: Path) -> None:
        """excel.info returns a dict of workbook metadata."""
        from otutil.tools.excel import info

        result = info(filepath=str(tmp_excel))
        assert isinstance(result, dict), f"Expected dict, got {type(result)}"
        assert "sheets" in result

    def test_sheets_returns_list(self, tmp_excel: Path) -> None:
        """excel.sheets returns a list of sheet dicts."""
        from otutil.tools.excel import sheets

        result = sheets(filepath=str(tmp_excel))
        assert isinstance(result, list), f"Expected list, got {type(result)}"
        assert len(result) >= 1

    def test_search_returns_list(self, tmp_excel: Path) -> None:
        """excel.search returns a list of matches."""
        from otutil.tools.excel import search

        result = search(filepath=str(tmp_excel), pattern="foo")
        assert isinstance(result, list), f"Expected list, got {type(result)}"
