"""Unit tests for Excel tool.

Tests excel.create(), excel.read(), excel.write(), etc.
Uses tmp_path fixture for isolated test files.
"""

from __future__ import annotations

import json
from typing import TYPE_CHECKING, Any

import pytest

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl not installed (install onetool-mcp[util])")

if TYPE_CHECKING:
    from pathlib import Path


def _to_str(result: Any) -> str:
    """Convert result to string for assertion comparisons.

    Since tools now return native types, we convert to JSON string
    for tests that check string presence.
    """
    if isinstance(result, str):
        return result
    return json.dumps(result)


@pytest.fixture
def excel_file(tmp_path: Path) -> Path:
    """Return path to a temp Excel file."""
    return tmp_path / "test.xlsx"


@pytest.mark.unit
@pytest.mark.tools
def test_pack_is_excel() -> None:
    """Verify pack is correctly set."""
    from otutil.tools.excel import pack

    assert pack == "excel"


@pytest.mark.unit
@pytest.mark.tools
def test_all_exports_only_public_functions() -> None:
    """Verify __all__ contains the expected public functions."""
    from otutil.tools.excel import __all__

    expected = {
        # Core operations
        "create",
        "add_sheet",
        "read",
        "write",
        "info",
        "formula",
        # Range manipulation
        "cell_range",
        "cell_shift",
        # Search
        "search",
        # Table operations
        "tables",
        "table_info",
        "table_data",
        "create_table",
        # Structure manipulation
        "insert_rows",
        "delete_rows",
        "insert_cols",
        "delete_cols",
        "copy_range",
        # Extended inspection
        "sheets",
        "used_range",
        "formulas",
        "hyperlinks",
        "merged_cells",
        "named_ranges",
    }
    assert set(__all__) == expected


@pytest.mark.unit
@pytest.mark.tools
def test_create_workbook(excel_file: Path) -> None:
    """Verify excel.create() creates a new workbook."""
    from otutil.tools.excel import create

    result = create(filepath=str(excel_file))

    assert "Created workbook" in result
    assert excel_file.exists()


@pytest.mark.unit
@pytest.mark.tools
def test_create_workbook_with_sheet_name(excel_file: Path) -> None:
    """Verify excel.create() creates workbook with custom sheet name."""
    from otutil.tools.excel import create, info

    create(filepath=str(excel_file), sheet_name="Sales")
    result = info(filepath=str(excel_file))

    assert "Sales" in _to_str(result)


@pytest.mark.unit
@pytest.mark.tools
def test_create_workbook_creates_parent_dirs(tmp_path: Path) -> None:
    """Verify excel.create() creates parent directories."""
    from otutil.tools.excel import create

    nested_file = tmp_path / "nested" / "dir" / "test.xlsx"
    result = create(filepath=str(nested_file))

    assert "Created workbook" in result
    assert nested_file.exists()


@pytest.mark.unit
@pytest.mark.tools
def test_add_sheet(excel_file: Path) -> None:
    """Verify excel.add_sheet() adds a new sheet."""
    from otutil.tools.excel import add_sheet, create, info

    create(filepath=str(excel_file))
    result = add_sheet(filepath=str(excel_file), sheet_name="Summary")

    assert "Created sheet" in result
    info_result = info(filepath=str(excel_file))
    assert "Summary" in _to_str(info_result)


@pytest.mark.unit
@pytest.mark.tools
def test_add_sheet_duplicate_error(excel_file: Path) -> None:
    """Verify excel.add_sheet() errors on duplicate sheet name."""
    from otutil.tools.excel import add_sheet, create

    create(filepath=str(excel_file), sheet_name="Data")
    result = add_sheet(filepath=str(excel_file), sheet_name="Data")

    assert "Error" in result
    assert "already exists" in result


@pytest.mark.unit
@pytest.mark.tools
def test_add_sheet_file_not_found() -> None:
    """Verify excel.add_sheet() returns error for missing file."""
    from otutil.tools.excel import add_sheet

    result = add_sheet(filepath="/nonexistent/path/missing.xlsx", sheet_name="Test")

    assert "Error" in result
    assert "not found" in result


@pytest.mark.unit
@pytest.mark.tools
def test_write_and_read(excel_file: Path) -> None:
    """Verify excel.write() and excel.read() round-trip data."""
    from otutil.tools.excel import create, read, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["Name", "Value"], ["A", 1], ["B", 2]])
    result = read(filepath=str(excel_file))

    result_str = _to_str(result)
    assert "Name" in result_str
    assert "Value" in result_str
    assert "A" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_write_to_specific_cell(excel_file: Path) -> None:
    """Verify excel.write() writes to specified start cell."""
    from otutil.tools.excel import create, write

    create(filepath=str(excel_file))
    result = write(
        filepath=str(excel_file),
        data=[["X", "Y"]],
        start_cell="C3",
    )

    assert "Wrote" in result


@pytest.mark.unit
@pytest.mark.tools
def test_write_to_specific_sheet(excel_file: Path) -> None:
    """Verify excel.write() writes to specified sheet."""
    from otutil.tools.excel import add_sheet, create, read, write

    create(filepath=str(excel_file))
    add_sheet(filepath=str(excel_file), sheet_name="Data")
    write(
        filepath=str(excel_file),
        data=[["Test", 123]],
        sheet_name="Data",
    )
    result = read(filepath=str(excel_file), sheet_name="Data")

    result_str = _to_str(result)
    assert "Test" in result_str
    assert "123" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_write_empty_data_returns_error(excel_file: Path) -> None:
    """Verify excel.write() returns error for empty data."""
    from otutil.tools.excel import create, write

    create(filepath=str(excel_file))
    result = write(filepath=str(excel_file), data=[])

    assert "Error" in result
    assert "No data" in result


@pytest.mark.unit
@pytest.mark.tools
def test_read_specific_range(excel_file: Path) -> None:
    """Verify excel.read() reads from specified range."""
    from otutil.tools.excel import create, read, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A", "B", "C"], [1, 2, 3], [4, 5, 6]])
    result = read(filepath=str(excel_file), start_cell="B1", end_cell="C2")

    result_str = _to_str(result)
    assert "B" in result_str
    assert "C" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_read_nonexistent_sheet_returns_error(excel_file: Path) -> None:
    """Verify excel.read() returns error for nonexistent sheet."""
    from otutil.tools.excel import create, read

    create(filepath=str(excel_file))
    result = read(filepath=str(excel_file), sheet_name="Missing")

    assert "Error" in result
    assert "not found" in result


@pytest.mark.unit
@pytest.mark.tools
def test_info_returns_metadata(excel_file: Path) -> None:
    """Verify excel.info() returns workbook metadata."""
    from otutil.tools.excel import create, info

    create(filepath=str(excel_file), sheet_name="Data")
    result = info(filepath=str(excel_file))

    result_str = _to_str(result)
    assert "test.xlsx" in result_str
    assert "Data" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_info_with_ranges(excel_file: Path) -> None:
    """Verify excel.info() includes used ranges when requested."""
    from otutil.tools.excel import create, info, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A", "B"], [1, 2], [3, 4]])
    result = info(filepath=str(excel_file), include_ranges=True)

    # Should show range like A1:B3
    result_str = _to_str(result)
    assert "A1" in result_str or "range" in result_str.lower()


@pytest.mark.unit
@pytest.mark.tools
def test_formula(excel_file: Path) -> None:
    """Verify excel.formula() applies formula to cell."""
    from otutil.tools.excel import create, formula, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["Value"], [10], [20], [30]])
    result = formula(filepath=str(excel_file), cell="A5", formula="=SUM(A2:A4)")

    assert "Applied formula" in result
    assert "SUM" in result


@pytest.mark.unit
@pytest.mark.tools
def test_formula_auto_prepends_equals(excel_file: Path) -> None:
    """Verify excel.formula() adds = prefix if missing."""
    from otutil.tools.excel import create, formula

    create(filepath=str(excel_file))
    result = formula(filepath=str(excel_file), cell="A1", formula="TODAY()")

    assert "=TODAY()" in result


@pytest.mark.unit
@pytest.mark.tools
def test_formula_to_specific_sheet(excel_file: Path) -> None:
    """Verify excel.formula() applies to specified sheet."""
    from otutil.tools.excel import add_sheet, create, formula

    create(filepath=str(excel_file))
    add_sheet(filepath=str(excel_file), sheet_name="Calcs")
    result = formula(
        filepath=str(excel_file),
        cell="A1",
        formula="=1+1",
        sheet_name="Calcs",
    )

    assert "Applied formula" in result


@pytest.mark.unit
@pytest.mark.tools
def test_read_nonexistent_file() -> None:
    """Verify excel.read() returns error for missing file."""
    from otutil.tools.excel import read

    result = read(filepath="/nonexistent/path/missing.xlsx")

    assert "Error" in result
    assert "not found" in result


@pytest.mark.unit
@pytest.mark.tools
def test_info_nonexistent_file() -> None:
    """Verify excel.info() returns error for missing file."""
    from otutil.tools.excel import info

    result = info(filepath="/nonexistent/path/missing.xlsx")

    assert "Error" in result
    assert "not found" in result


@pytest.mark.unit
@pytest.mark.tools
def test_write_nonexistent_file() -> None:
    """Verify excel.write() returns error for missing file."""
    from otutil.tools.excel import write

    result = write(filepath="/nonexistent/path/missing.xlsx", data=[["test"]])

    assert "Error" in result
    assert "not found" in result


@pytest.mark.unit
@pytest.mark.tools
def test_formula_nonexistent_file() -> None:
    """Verify excel.formula() returns error for missing file."""
    from otutil.tools.excel import formula

    result = formula(filepath="/nonexistent/path/missing.xlsx", cell="A1", formula="=1")

    assert "Error" in result
    assert "not found" in result


# =============================================================================
# Range Manipulation (Pure Functions)
# =============================================================================


@pytest.mark.unit
@pytest.mark.tools
def test_cell_range_expand() -> None:
    """Verify cell_range expands a cell into a range."""
    from otutil.tools.excel import cell_range

    result = cell_range(cell="A1", right=5, down=5)

    assert "A1:F6" in result


@pytest.mark.unit
@pytest.mark.tools
def test_cell_range_expand_left_up() -> None:
    """Verify cell_range can expand left and up."""
    from otutil.tools.excel import cell_range

    result = cell_range(cell="C3", left=2, up=2)

    assert "A1:C3" in result


@pytest.mark.unit
@pytest.mark.tools
def test_cell_shift() -> None:
    """Verify cell_shift moves a cell reference."""
    from otutil.tools.excel import cell_shift

    result = cell_shift(cell="A1", rows=5, cols=3)

    assert "D6" in result


@pytest.mark.unit
@pytest.mark.tools
def test_cell_shift_single_direction() -> None:
    """Verify cell_shift works with single direction."""
    from otutil.tools.excel import cell_shift

    result = cell_shift(cell="B2", rows=3)

    assert "B5" in result


# =============================================================================
# Search
# =============================================================================


@pytest.mark.unit
@pytest.mark.tools
def test_search_wildcard(excel_file: Path) -> None:
    """Verify search finds values with wildcard pattern."""
    from otutil.tools.excel import create, search, write

    create(filepath=str(excel_file))
    write(
        filepath=str(excel_file),
        data=[["Name"], ["Widget A"], ["Gadget B"], ["Widget C"]],
    )
    result = search(filepath=str(excel_file), pattern="Widget*")

    result_str = _to_str(result)
    assert "Widget A" in result_str
    assert "Widget C" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_search_first_only(excel_file: Path) -> None:
    """Verify search returns single-item list when first_only=True."""
    from otutil.tools.excel import create, search, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["Name"], ["Test1"], ["Test2"], ["Test3"]])
    result = search(filepath=str(excel_file), pattern="Test*", first_only=True)

    # first_only=True returns a JSON list with one item for API consistency
    import json

    data = json.loads(result)
    assert isinstance(data, list)
    assert len(data) == 1
    assert data[0]["value"] == "Test1"
    assert "cell" in data[0]


@pytest.mark.unit
@pytest.mark.tools
def test_search_regex(excel_file: Path) -> None:
    """Verify search works with regex patterns."""
    from otutil.tools.excel import create, search, write

    create(filepath=str(excel_file))
    write(
        filepath=str(excel_file), data=[["ID"], ["ABC-123"], ["XYZ-456"], ["ABC-789"]]
    )
    result = search(filepath=str(excel_file), pattern="^ABC", regex=True)

    result_str = _to_str(result)
    assert "ABC-123" in result_str
    assert "ABC-789" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_search_no_matches(excel_file: Path) -> None:
    """Verify search returns empty list when no matches."""
    from otutil.tools.excel import create, search, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["Name"], ["Alpha"], ["Beta"]])
    result = search(filepath=str(excel_file), pattern="Gamma*")

    # Result is a JSON string of an empty list
    import json

    assert result == "[]"


# =============================================================================
# Table Operations
# =============================================================================


@pytest.mark.unit
@pytest.mark.tools
def test_create_table(excel_file: Path) -> None:
    """Verify create_table creates a native Excel table."""
    from otutil.tools.excel import create, create_table, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["Name", "Value"], ["A", 100], ["B", 200]])
    result = create_table(
        filepath=str(excel_file), data_range="A1:B3", table_name="TestTable"
    )

    assert "Created table" in result
    assert "TestTable" in result


@pytest.mark.unit
@pytest.mark.tools
def test_create_table_auto_name(excel_file: Path) -> None:
    """Verify create_table auto-generates table name."""
    from otutil.tools.excel import create, create_table, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["X", "Y"], [1, 2]])
    result = create_table(filepath=str(excel_file), data_range="A1:B2")

    assert "Created table" in result
    assert "Table1" in result


@pytest.mark.unit
@pytest.mark.tools
def test_tables_list(excel_file: Path) -> None:
    """Verify tables lists all tables in worksheet."""
    from otutil.tools.excel import create, create_table, tables, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A", "B"], [1, 2], [3, 4]])
    create_table(filepath=str(excel_file), data_range="A1:B3", table_name="MyTable")
    result = tables(filepath=str(excel_file))

    result_str = _to_str(result)
    assert "MyTable" in result_str
    assert "A1:B3" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_table_info(excel_file: Path) -> None:
    """Verify table_info returns detailed table information."""
    from otutil.tools.excel import create, create_table, table_info, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["Col1", "Col2"], ["X", 1], ["Y", 2]])
    create_table(filepath=str(excel_file), data_range="A1:B3", table_name="InfoTable")
    result = table_info(filepath=str(excel_file), table_name="InfoTable")

    result_str = _to_str(result)
    assert "InfoTable" in result_str
    assert "Col1" in result_str
    assert "Col2" in result_str
    assert "row_count" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_table_data(excel_file: Path) -> None:
    """Verify table_data returns table contents as dicts."""
    from otutil.tools.excel import create, create_table, table_data, write

    create(filepath=str(excel_file))
    write(
        filepath=str(excel_file), data=[["Name", "Score"], ["Alice", 95], ["Bob", 87]]
    )
    create_table(filepath=str(excel_file), data_range="A1:B3", table_name="DataTable")
    result = table_data(filepath=str(excel_file), table_name="DataTable")

    result_str = _to_str(result)
    assert "Alice" in result_str
    assert "Bob" in result_str
    assert "95" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_table_data_single_row(excel_file: Path) -> None:
    """Verify table_data returns single row by index."""
    from otutil.tools.excel import create, create_table, table_data, write

    create(filepath=str(excel_file))
    write(
        filepath=str(excel_file), data=[["Name", "Score"], ["Alice", 95], ["Bob", 87]]
    )
    create_table(filepath=str(excel_file), data_range="A1:B3", table_name="RowTable")
    result = table_data(filepath=str(excel_file), table_name="RowTable", row_index=0)

    result_str = _to_str(result)
    assert "Alice" in result_str
    assert "Bob" not in result_str


# =============================================================================
# Structure Manipulation
# =============================================================================


@pytest.mark.unit
@pytest.mark.tools
def test_insert_rows(excel_file: Path) -> None:
    """Verify insert_rows adds rows at specified position."""
    from otutil.tools.excel import create, insert_rows, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A"], [1], [2]])
    result = insert_rows(filepath=str(excel_file), row=2, count=2)

    assert "Inserted 2 rows" in result


@pytest.mark.unit
@pytest.mark.tools
def test_delete_rows(excel_file: Path) -> None:
    """Verify delete_rows removes rows at specified position."""
    from otutil.tools.excel import create, delete_rows, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A"], [1], [2], [3], [4]])
    result = delete_rows(filepath=str(excel_file), row=2, count=2)

    assert "Deleted 2 rows" in result


@pytest.mark.unit
@pytest.mark.tools
def test_insert_cols(excel_file: Path) -> None:
    """Verify insert_cols adds columns at specified position."""
    from otutil.tools.excel import create, insert_cols, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A", "B", "C"]])
    result = insert_cols(filepath=str(excel_file), col="B", count=2)

    assert "Inserted 2 columns" in result


@pytest.mark.unit
@pytest.mark.tools
def test_insert_cols_by_number(excel_file: Path) -> None:
    """Verify insert_cols works with column number."""
    from otutil.tools.excel import create, insert_cols, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A", "B", "C"]])
    result = insert_cols(filepath=str(excel_file), col=2, count=1)

    assert "Inserted 1 column" in result
    assert "columns" not in result  # Singular form


@pytest.mark.unit
@pytest.mark.tools
def test_write_singular_row_grammar(excel_file: Path) -> None:
    """Verify write uses singular 'row' for count of 1."""
    from otutil.tools.excel import create, write

    create(filepath=str(excel_file))
    result = write(filepath=str(excel_file), data=[["Single"]])

    assert "Wrote 1 row" in result
    assert "rows" not in result


@pytest.mark.unit
@pytest.mark.tools
def test_insert_rows_singular_grammar(excel_file: Path) -> None:
    """Verify insert_rows uses singular 'row' for count of 1."""
    from otutil.tools.excel import create, insert_rows

    create(filepath=str(excel_file))
    result = insert_rows(filepath=str(excel_file), row=1, count=1)

    assert "Inserted 1 row" in result
    assert "rows" not in result


@pytest.mark.unit
@pytest.mark.tools
def test_delete_rows_singular_grammar(excel_file: Path) -> None:
    """Verify delete_rows uses singular 'row' for count of 1."""
    from otutil.tools.excel import create, delete_rows, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A"], ["B"], ["C"]])
    result = delete_rows(filepath=str(excel_file), row=2, count=1)

    assert "Deleted 1 row" in result
    assert "rows" not in result


@pytest.mark.unit
@pytest.mark.tools
def test_delete_cols_singular_grammar(excel_file: Path) -> None:
    """Verify delete_cols uses singular 'column' for count of 1."""
    from otutil.tools.excel import create, delete_cols, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A", "B", "C"]])
    result = delete_cols(filepath=str(excel_file), col="B", count=1)

    assert "Deleted 1 column" in result
    assert "columns" not in result


@pytest.mark.unit
@pytest.mark.tools
def test_write_create_if_missing(tmp_path: Path) -> None:
    """Verify write creates file when create_if_missing=True."""
    from otutil.tools.excel import read, write

    new_file = tmp_path / "new_workbook.xlsx"
    result = write(
        filepath=str(new_file),
        data=[["Header"], ["Value"]],
        create_if_missing=True,
    )

    assert "Wrote 2 rows" in result
    assert new_file.exists()
    # Verify data was written
    data = read(filepath=str(new_file))
    data_str = _to_str(data)
    assert "Header" in data_str
    assert "Value" in data_str


@pytest.mark.unit
@pytest.mark.tools
def test_write_create_if_missing_with_sheet_name(tmp_path: Path) -> None:
    """Verify write creates file with custom sheet name when create_if_missing=True."""
    from otutil.tools.excel import info, write

    new_file = tmp_path / "new_with_sheet.xlsx"
    write(
        filepath=str(new_file),
        data=[["Test"]],
        sheet_name="CustomSheet",
        create_if_missing=True,
    )

    result = info(filepath=str(new_file))
    result_str = _to_str(result)
    assert "CustomSheet" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_write_create_if_missing_creates_parent_dirs(tmp_path: Path) -> None:
    """Verify write creates parent directories when create_if_missing=True."""
    from otutil.tools.excel import write

    nested_file = tmp_path / "nested" / "dir" / "new.xlsx"
    result = write(
        filepath=str(nested_file),
        data=[["Data"]],
        create_if_missing=True,
    )

    assert "Wrote 1 row" in result
    assert nested_file.exists()


@pytest.mark.unit
@pytest.mark.tools
def test_delete_cols(excel_file: Path) -> None:
    """Verify delete_cols removes columns at specified position."""
    from otutil.tools.excel import create, delete_cols, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A", "B", "C", "D", "E"]])
    result = delete_cols(filepath=str(excel_file), col="C", count=2)

    assert "Deleted 2 columns" in result


@pytest.mark.unit
@pytest.mark.tools
def test_copy_range(excel_file: Path) -> None:
    """Verify copy_range copies cells to new location."""
    from otutil.tools.excel import copy_range, create, read, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["X", "Y"], [1, 2], [3, 4]])
    result = copy_range(filepath=str(excel_file), source="A1:B3", target="D1")

    assert "Copied" in result
    # Verify the data was copied
    data = read(filepath=str(excel_file))
    data_str = _to_str(data)
    assert "X" in data_str  # Original
    assert "D1:E3" in result or "Copied A1:B3" in result


@pytest.mark.unit
@pytest.mark.tools
def test_copy_range_to_different_sheet(excel_file: Path) -> None:
    """Verify copy_range can copy to a different sheet."""
    from otutil.tools.excel import add_sheet, copy_range, create, read, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["Data"], [100], [200]])
    add_sheet(filepath=str(excel_file), sheet_name="Backup")
    result = copy_range(
        filepath=str(excel_file), source="A1:A3", target="A1", target_sheet="Backup"
    )

    assert "Copied" in result
    backup_data = read(filepath=str(excel_file), sheet_name="Backup")
    backup_str = _to_str(backup_data)
    assert "Data" in backup_str
    assert "100" in backup_str


# =============================================================================
# Extended Inspection
# =============================================================================


@pytest.mark.unit
@pytest.mark.tools
def test_sheets(excel_file: Path) -> None:
    """Verify sheets lists all sheets with visibility."""
    from otutil.tools.excel import add_sheet, create, sheets

    create(filepath=str(excel_file), sheet_name="Main")
    add_sheet(filepath=str(excel_file), sheet_name="Data")
    add_sheet(filepath=str(excel_file), sheet_name="Summary")
    result = sheets(filepath=str(excel_file))

    result_str = _to_str(result)
    assert "Main" in result_str
    assert "Data" in result_str
    assert "Summary" in result_str
    assert "visible" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_used_range(excel_file: Path) -> None:
    """Verify used_range returns the data extent."""
    from otutil.tools.excel import create, used_range, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["A", "B", "C"], [1, 2, 3], [4, 5, 6]])
    result = used_range(filepath=str(excel_file))

    assert "A1:C3" in result


@pytest.mark.unit
@pytest.mark.tools
def test_used_range_minimal(excel_file: Path) -> None:
    """Verify used_range returns minimal range for new sheet.

    Note: openpyxl reports A1:A1 for empty sheets (max_row=1, max_column=1).
    """
    from otutil.tools.excel import create, used_range

    create(filepath=str(excel_file))
    result = used_range(filepath=str(excel_file))

    # New sheets report A1:A1 as the used range
    assert result == "A1:A1"


@pytest.mark.unit
@pytest.mark.tools
def test_formulas_list(excel_file: Path) -> None:
    """Verify formulas lists all formula cells."""
    from otutil.tools.excel import create, formula, formulas, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["Value"], [10], [20], [30]])
    formula(filepath=str(excel_file), cell="A5", formula="=SUM(A2:A4)")
    formula(filepath=str(excel_file), cell="A6", formula="=AVERAGE(A2:A4)")
    result = formulas(filepath=str(excel_file))

    result_str = _to_str(result)
    assert "SUM" in result_str
    assert "AVERAGE" in result_str
    assert "A5" in result_str
    assert "A6" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_merged_cells(excel_file: Path) -> None:
    """Verify merged_cells lists merged ranges."""
    # Create workbook and merge cells using openpyxl directly
    from openpyxl import load_workbook

    from otutil.tools.excel import create, merged_cells

    create(filepath=str(excel_file))
    wb = load_workbook(str(excel_file))
    ws = wb.active
    ws.merge_cells("A1:C1")
    ws.merge_cells("B3:D5")
    wb.save(str(excel_file))

    result = merged_cells(filepath=str(excel_file))

    result_str = _to_str(result)
    assert "A1:C1" in result_str
    assert "B3:D5" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_named_ranges(excel_file: Path) -> None:
    """Verify named_ranges lists defined names."""
    # Create workbook with named range using openpyxl directly
    from openpyxl import load_workbook
    from openpyxl.workbook.defined_name import DefinedName

    from otutil.tools.excel import create, named_ranges, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["Sales"], [100], [200], [300]])
    wb = load_workbook(str(excel_file))
    # Add a named range
    defn = DefinedName("SalesData", attr_text="Sheet1!$A$1:$A$4")
    wb.defined_names.add(defn)
    wb.save(str(excel_file))

    result = named_ranges(filepath=str(excel_file))

    result_str = _to_str(result)
    assert "SalesData" in result_str


@pytest.mark.unit
@pytest.mark.tools
def test_hyperlinks_empty(excel_file: Path) -> None:
    """Verify hyperlinks returns empty list when no links."""
    from otutil.tools.excel import create, hyperlinks, write

    create(filepath=str(excel_file))
    write(filepath=str(excel_file), data=[["Text"], ["Hello"], ["World"]])
    result = hyperlinks(filepath=str(excel_file))

    # Result is a JSON string of an empty list
    assert result == "[]"
